# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from .basex import SheetBase, BookBase
from .writermixin import SheetMixin, BookMixin
from .utils import tag_test, parse_cell_tag, recompile_patterns
from .xlnode import Tree, Row, Cell, EmptyCell, Node, create_cell
from .jinja import JinjaEnvx
from .nodemap import NodeMap
from .sheetresource import SheetResource
from .richtexthandler import rich_handlerx
from .mergerx import Merger
from .config import config
from .misc import CellTag

class SheetWriter(SheetBase, SheetMixin):

    def __init__(self, bookwriter, sheet_resource, sheet_name):
        self.workbook = bookwriter.workbook
        self.merger = sheet_resource.merger
        self.rdsheet = sheet_resource.rdsheet
        self.wtsheet = self.workbook.create_sheet(title=sheet_name)
        self.copy_sheet_settings()
        self.wtrows = set()
        self.wtcols = set()
        self.min_rowx = 0
        self.min_colx = 0
        self.reset_pos()


class BookWriter(BookBase, BookMixin):
    sheet_writer_cls = SheetWriter

    def __init__(self, fname, debug=False, user_extensions=None):
        if user_extensions is None:
            user_extensions = []
        config.debug = debug
        self.user_extensions = user_extensions
        self.load(fname)

    def load(self, fname):
        self.workbook = load_workbook(fname)
        self.font_map = {}
        self.sheet_writer_map = {}
        self.sheet_resource_map = {}
        self.node_map = NodeMap()
        self.jinja_env = JinjaEnvx(self.node_map, user_extensions=self.user_extensions)
        for index,rdsheet in enumerate(self.workbook.worksheets):
            merger = Merger(rdsheet)
            sheet_tree = self.build(rdsheet, index, merger)
            sheet_resource = SheetResource(rdsheet, sheet_tree, self.jinja_env, merger)
            self.put_sheet_resource(index, rdsheet.title, sheet_resource)
            self.workbook.remove(rdsheet)

    def build(self, sheet, index, merger):
        tree = Tree(index, self.node_map)
        user_tags = [ tag for ext in self.user_extensions for tag in ext.tags] #flattening
        recompile_patterns(user_tags)
        max_row = max(sheet.max_row, merger.image_merger.max_row)
        max_col = max(sheet.max_column, merger.image_merger.max_col)
        for rowx in range(1, max_row + 1):
            row_node = Row(rowx)
            tree.add_child(row_node)
            for colx in range(1, max_col + 1):
                sheet_cell = sheet._cells.get((rowx, colx))
                if not sheet_cell:
                    cell_node = EmptyCell(rowx, colx)
                    tree.add_child(cell_node)
                    continue
                cell_tag_map = None
                if sheet_cell.comment:
                    comment = sheet_cell.comment.text
                    if tag_test(comment):
                        _,cell_tag_map = parse_cell_tag(comment)
                value = sheet_cell._value
                data_type = sheet_cell.data_type
                rich_text = None
                if hasattr(value, 'rich') and value.rich:
                    rich_text = value.rich
                if data_type == 's':
                    if not tag_test(value):
                        cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                    else:
                        font = self.get_font(sheet_cell._style.fontId)
                        cell_node = create_cell(sheet_cell, rowx, colx, value, rich_text, data_type, font, rich_handlerx, user_tags)
                else:
                    cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                if cell_tag_map:
                    cell_tag = CellTag(cell_tag_map)
                    cell_node.extend_cell_tag(cell_tag)
                    if colx==1:
                        row_node.cell_tag = cell_tag
                tree.add_child(cell_node)
            tree.add_child(Node())#
        return tree


    def save(self, fname):
        if not self.workbook.active:
            self.workbook.active = 0
        self.workbook.save(fname)
        for sheet in self.workbook.worksheets:
            self.workbook.remove(sheet)
        self.sheet_writer_map.clear()
